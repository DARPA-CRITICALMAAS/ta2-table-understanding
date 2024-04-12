from pathlib import Path

import openpyxl
import pandas as pd
import serde.csv
from tqdm.auto import tqdm

(infile,) = Path(__file__).parent.glob("*.xlsx")

wb = openpyxl.load_workbook(infile, data_only=True)
ws = wb.worksheets[0]
rows = list(ws.rows)
records = []
for i in tqdm(range(6, 2306 + 1)):
    record = {}

    for C in ["A", "B"]:
        record[ws[f"{C}5"].value] = ws[f"{C}{i}"].value

    for C in ["BL", "BM", "BN"]:
        record[" ".join(ws[f"{C}{j}"].value for j in range(4, 6))] = ws[f"{C}{i}"].value

    for C in ["F", "AJ", "G", "AK"]:
        key = " ".join(ws[f"{C}{j}"].value for j in range(4, 6))
        if C == "G":
            key = "Min. ResC " + key
        elif C == "AK":
            key = "Min. ResV " + key
        record[key] = ws[f"{C}{i}"].value

    records.append(record)

df = pd.DataFrame(records)
df.to_csv(
    Path(__file__).parent.parent.parent.parent
    / "examples/tables"
    / (infile.stem + ".csv"),
    index=False,
)

# cols = ['Deposit/Project Name', 'Country']
# df[df[cols].apply(lambda row: '__'.join(row.values.astype(str)), axis=1).duplicated()]
