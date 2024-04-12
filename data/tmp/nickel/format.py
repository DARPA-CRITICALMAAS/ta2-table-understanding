from pathlib import Path

import openpyxl
import pandas as pd
import serde.csv
from tqdm.auto import tqdm

(infile,) = Path(__file__).parent.glob("*.xlsx")

wb = openpyxl.load_workbook(infile, data_only=True)
ws = wb.worksheets[0]
ws["K1"]
rows = list(ws.rows)
records = []
for i in tqdm(range(5, len(rows) + 1)):
    base_record = {
        ws["B2"].value: ws[f"B{i}"].value,
        ws["C2"].value: ws[f"C{i}"].value,
        ws["E3"].value: ws[f"E{i}"].value,
        ws["F3"].value: ws[f"F{i}"].value,
        ws["G2"].value: ws[f"H{i}"].value,
        ws["DP2"].value: ws[f"DP{i}"].value,
        # code compliance
        ws["DL1"].value: ws[f"DL{i}"].value,
        ws["DM1"].value: ws[f"DM{i}"].value,
        " ".join(ws[f"FR{j}"].value for j in range(2, 5)): ws[f"FR{i}"].value,
        " ".join(ws[f"FS{j}"].value for j in range(2, 5)): ws[f"FS{i}"].value,
        " ".join(ws[f"FT{j}"].value for j in range(2, 5)): ws[f"FT{i}"].value,
    }

    if (ws[f"C{i}"].value or "").strip() == "":
        continue

    kidx = ws[f"K1"].col_idx
    diff = 21
    for k in range(5):
        record = base_record.copy()
        record["Category"] = ws.cell(row=3, column=kidx + k * diff).value
        has_data = {}
        for j in ["K", "S", "T"]:
            col_idx = ws[f"{j}1"].col_idx + k * diff
            val = ws.cell(row=i, column=col_idx).value
            record[ws.cell(row=4, column=col_idx).value] = val
            has_data[j] = val is not None

        if has_data["S"] or has_data["T"]:
            records.append(record)


pd.DataFrame(records).to_csv(
    Path(__file__).parent.parent.parent.parent
    / "examples/tables"
    / (infile.stem + ".csv"),
    index=False,
)
